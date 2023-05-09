import os
from openpyxl import load_workbook
from AbstractParser import AbstractParser

# openpyxl从1开始取值

class ExcelParser(AbstractParser):
    """解析excel文件"""
    def __init__(self,excelPath):
        self.data_path = excelPath
        # 获取到excel指定对象
        wb = load_workbook(self.data_path)
        # 只取第一张表
        sheet_name = wb.sheetnames[0]
        self.sheet = wb[sheet_name]
    
    def get_header_row(self):
        return self.get_row_value(1)
        
    """获取某一行的数据"""
    def get_row_value(self, row_num):
        row = row_num + 2
        row_value_list = []
        for y in range(2, self.sheet.max_column + 1):
            value = self.sheet.cell(row, y).value
            row_value_list.append(value)
        return row_value_list

    """获取某一列的数据"""
    def get_column_value(self, col_num):
        col_value_list = []
        for x in range(2, self.sheet.max_row + 1):
            value = self.sheet.cell(x, col_num).value
            col_value_list.append(value)
        return col_value_list

    # """获取某一个单元格的数据"""
    # def get_cell_value(self, raw_no, col_no):
    #     sh = self.wb[self.sheet_name]
    #     value = sh.cell(raw_no, col_no).value
    #     return value

    # """向某个单元格写入数据"""
    # def write_cell(self, raw_no, col_no, value):
    #     sh = self.wb[self.sheet_name]
    #     sh.cell(raw_no, col_no).value = value
    #     self.wb.save(self.data_path)

if __name__ == '__main__':
    workspace_path = os.path.dirname(__file__)
    # excel_path = workspace_path+"/data.csv"
    excel_path = "/Users/lettyliu/Git/AITranslator/AITranslator/csv_demo.csv"
    parser = ExcelParser(excel_path)
    print(parser.get_cell_value(1,1))
