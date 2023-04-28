# import xlrd
import sys
from openpyxl import load_workbook

class ExcelParser(object):
    """解析excel文件"""
    def __init__(self,excelPath):
        self.data_path = excelPath
        # 获取到excel指定对象
        self.wb = load_workbook(self.data_path)
        name_list = excelPath.split('/')[-1].split('_')
        # self.sheet_name = name_list[0] + "_" + name_list[1]
        self.sheet_name = self.wb.sheetnames[0]
        
    def get_row_value(self, raw_no):
        """获取某一行的数据"""
        sh = self.wb[self.sheet_name]
        row_value_list = []
        for y in range(2, sh.max_column + 1):
            value = sh.cell(raw_no, y).value
            row_value_list.append(value)
        return row_value_list

    def get_column_value(self, col_no):
        """获取某一列的数据"""
        sh = self.wb[self.sheet_name]
        col_value_list = []
        for x in range(2, sh.max_row + 1):
            value = sh.cell(x, col_no).value
            col_value_list.append(value)
        return col_value_list

    def get_cell_value(self, raw_no, col_no):
        """获取某一个单元格的数据"""
        sh = self.wb[self.sheet_name]
        value = sh.cell(raw_no, col_no).value
        return value

    def write_cell(self, raw_no, col_no, value):
        """向某个单元格写入数据"""
        sh = self.wb[self.sheet_name]
        sh.cell(raw_no, col_no).value = value
        self.wb.save(self.data_path)
